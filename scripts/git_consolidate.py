import csv
import os
import sys

from openpyxl import Workbook, load_workbook


def parse_folder(folder, target, file_data):


    for (dirpath, dirnames, filenames) in os.walk(folder):
        for filename in filenames:
            #print('file_data[\'{}\'] = {{\'app_name\': \'\'}}'.format(filename))
            if file_data.get(filename):
                print(file_data[filename]['app_name'])
                source = os.path.join(dirpath, filename)
                process_file(source, target, ['hash', 'name', 'e-mail', 'date', 'description', 'app'],
                             app_name=file_data[filename]['app_name'])


def process_file(source_file, target_file, headers, **kwargs):
    if os.path.exists(target_file):
        wb = load_workbook(filename=target_file, data_only=True)
        sheet = wb['Commits']
        row_num = sheet.max_row + 1
    else:
        wb = Workbook()
        sheet = wb.create_sheet('Commits')
        row_num = 1
        col = 1
        for header in headers:
            sheet.cell(row=row_num, column=col, value=header)
            col += 1
        row_num += 1
    commit_count = 0
    with open(source_file, 'r', encoding='utf-8') as pike_file:
        reader = csv.reader(pike_file, delimiter='|')
        for row in reader:
            for i in range(len(row)):
                sheet.cell(row=row_num, column=i + 1, value=row[i])
            sheet.cell(row=row_num, column=6, value=kwargs.get('app_name'))
            commit_count += 1
            row_num += 1
    print('\tWrote {} for App {} to {}'.format(commit_count, kwargs.get('app_name'), target_file))
    wb.save(target_file)


if __name__ == '__main__':
    """
    Dentro del repositorio local del proyecto (donde está el directorio .git):
    Corre el siguiente commando::
    
        $ git log --pretty=tformat:"%h|%an|%ae|%ai|%s" > nombre_proyecto.pike
    
    El nombre del archivo debe coincidir con el nombre registrado en la variable  file_data donde se asocia
    el nombre del proyecto con el nombre del archivo.
    
    Copia todos los archivos .pike al directorio ./output/git/pikes:
    
    Activa un ambiente virtual e instala los requerimiento:
        pip install openpyxl
        
    No corre dentro del contenedor. Se debe correr desde lines de comando
        
    Correr 
    
        $ python git_consolidate.py git_commits.xlsx
  
    """


    file_data = dict()


    # Black Box
    file_data['blackbox-api.pike'] = {'app_name': 'Black Box'}
    file_data['blackbox-ui.pike'] = {'app_name': 'Black Box'}
    file_data['blackbox_raspberry.pike'] = {'app_name': 'Black Box'}
    # Chat Bot Marine Traffic
    file_data['bomarinetraffic.pike'] = {'app_name': 'ChatBot'}

    # Scientific Apps
    file_data['ctan_display.pike'] = {'app_name': 'CTAN'}
    file_data['signalStation.pike'] = {'app_name': 'Signal Station'}
    file_data['harborWorkbench.pike'] = {'app_name': 'Harbor Workbench'}
    file_data['lockages_managment.pike'] = {'app_name': 'Lockage Management System'}

    file_data['ais_app_i2.pike'] = {'app_name': 'AIS Server v3.0'}
    file_data['navAidsV2.pike'] = {'app_name': 'Navigation Aids Framework 3'}
    file_data['pilotAssigment.pike'] = {'app_name': 'Pilot Assignment'}
    file_data['SCIENTIFIC_APP_FWK_2.pike'] = {'app_name': 'Aplicaciones Cientificas Framework 2'}

    file_data['taf_displays_web_api.pike'] = {'app_name': 'TAF Display Web'}
    file_data['vesselScheduleWebApi.pike'] = {'app_name': 'Vessel Schedule Framework Web API'}
    
    file_data['TINO_APP_FWK_3.pike'] = {'app_name': 'Aplicaciones Cientificas Framework 3'}
    file_data['vessel_display_web.pike'] = {'app_name': 'Vessel Diaplay Web'}
    file_data['VESSEL_SCHED_APP_I2.pike'] = {'app_name': 'Vessel Scheduling WB v3.0'}
    file_data['vessel_schedulingCliente.pike'] = {'app_name': 'Vessel Scheduling WB v3.0'}
    file_data['vesselScheduleV1.pike'] = {'app_name': 'Vessel Schedule WB v1.0'}

    file_data['frameWork3Vessel.pike'] = {'app_name': 'Vessel Scheduling WB v3.0'}
    file_data['incidentDashboard.pike'] = {'app_name': 'Incident Dashboard'}
    file_data['INTER_APP_RELATION_I2.pike'] = {'app_name': 'Aplicaciones Cientificas Framework 3'}
    # PPU
    file_data['inventory_ppuApp.pike'] = {'app_name': 'PPU'}

    # Scientific Apps Bamboo
    file_data['wipVessel2.pike'] = {'app_name': 'Data Migration - WIP Bamboo'}
    file_data['wipVesselWeb.pike'] = {'app_name': 'Scripts de Bamboo'}

    # Scrips
    file_data['databasemigrations.pike'] = {'app_name': 'Data Migration - WIP Bamboo'}
    file_data['tino_scripts.pike'] = {'app_name': 'Scripts de Bamboo'}


    # VUMPA
    file_data['VUMPA_ADF.pike'] = {'app_name': 'VUMPA'}
    file_data['VUMPA_ADF_REPORTES.pike'] = {'app_name': 'VUMPA'}
    file_data['VUMPA_ADM_TABLES.pike'] = {'app_name': 'VUMPA'}
    file_data['VUMPA_BA.pike'] = {'app_name': 'VUMPA'}
    file_data['VUMPA_BPEL.pike'] = {'app_name': 'VUMPA'}
    file_data['VUMPA_DB.pike'] = {'app_name': 'VUMPA'}
    file_data['VUMPA_OSB.pike'] = {'app_name': 'VUMPA'}
    file_data['VUMPA_ZARPE.pike'] = {'app_name': 'VUMPA'}

    # Respaldo Financiero
    file_data['acpcommon.pike'] = {'app_name': 'Respaldo Financiero'}
    file_data['ProformaDB.pike'] = {'app_name': 'Respaldo Financiero'}
    file_data['respaldofinancieroframework.pike'] = {'app_name': 'Respaldo Financiero'}
    file_data['CSMSDBObject.pike'] = {'app_name': 'Respaldo Financiero'}
    file_data['respaldofinanciero.pike'] = {'app_name': 'Respaldo Financiero'}
    file_data['respaldofinancieroservices.pike'] = {'app_name': 'Respaldo Financiero'}
    file_data['EVTMSDBObjects.pike'] = {'app_name': 'Respaldo Financiero'}
    file_data['respaldofinancierodb.pike'] = {'app_name': 'Respaldo Financiero'}
    file_data['proforma.pike'] = {'app_name': 'Respaldo Financiero'}

    # Billing
    file_data['Billing_database_objects.pike'] = {'app_name': 'Billing'}
    file_data['Billing_Report.pike'] = {'app_name': 'Billing'}
    file_data['Billing.pike'] = {'app_name': 'Billing'}


    directory = r'.' ##.\output\git
    output_filename = sys.argv[1:][0] #'git_commits_ists-t.xlsx'
    target_filename = os.path.join(directory, output_filename)

    directory = r'.' #.\output\git\pikes
    parse_folder(directory, target_filename, file_data)
