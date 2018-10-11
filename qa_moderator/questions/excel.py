from openpyxl import Workbook

from qa_moderator.questions.models import Question


class QuestionExcelAdapter(object):

    def convert_all(self, filename):
        wb = Workbook()
        #sheet = wb.active
        sheet = wb.create_sheet(title='RCI Questions')
        questions = Question.objects.all()
        start_row = 1
        row = 1
        for question in questions:
            col = 1
            sheet.cell(column=col, row=row, value=question.question)
            col += 1
            sheet.cell(column=col, row=row, value=question.approved)
            col += 1
            sheet.cell(column=col, row=row, value=question.priority)
            col += 1
            sheet.cell(column=col, row=row, value=question.moderator_num)
            col += 1
            sheet.cell(column=col, row=row, value=question.created)
            row += 1
        wb.save(filename)
